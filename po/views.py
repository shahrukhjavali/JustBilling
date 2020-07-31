from django.shortcuts import render, redirect
from .models import Po, poItems, potracker
from products.models import Products
from vendor.models import Vendor
from master.models import UOM, Tax
from django.core.paginator import Paginator
from django.views.generic import View
from django.utils import timezone
from openpyxl import Workbook
from django.core.mail import EmailMessage
import random


class addPo(View):
    def get(self, request):
        venlist = Vendor.objects.all()
        podetails = Po.objects.filter(id=1)
        products = Products.objects.all()
        objs = UOM.objects.all()
        taxs = Tax.objects.all()
        poitms = poItems.objects.filter(podetails=podetails[0]).filter(ponum=podetails[0].ponum)
        return render(self.request, 'po/addpo.html', {'venlist': venlist,
                                                      'podetails': podetails, 'products': products, 'uomobjs': objs,
                                                      'poitms': poitms, 'taxobjs': taxs})

    def post(self, request):
        obj = Po()
        obj.ponum = str(getponum())
        obj.date = timezone.now()
        obj.vendor = Vendor.objects.get(id=1)
        obj.shipto = request.POST.get('adderss1') + request.POST.get('adderss2')
        obj.shipcity = request.POST.get('city')
        obj.shipstate = request.POST.get('state')
        obj.shippincode = request.POST.get('pincode')
        obj.poreqbydate = request.POST.get('poreq_date')
        obj.createdby = request.user
        obj.creation_date = timezone.now()
        obj.save()
        return render(self.request, 'po/addpo.html', {'podetails': obj})


def po(request):
    object = Po.objects.all().order_by("-id")
    paginator = Paginator(object, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'po/polist.html', {'po': page_obj})


def addchildItems(request, id):
    if request.method == 'POST':
        po = Po.objects.get(id=id)
        ponum = po.ponum
        product = Products.objects.get(id=int(request.POST.get('product')))
        qty = float(request.POST.get('qty'))
        uom = UOM.objects.get(name=request.POST.get('uom'))
        subtotal = qty * product.price
        obj = poItems()
        obj.podetails = po
        obj.ponum = ponum
        obj.po_items = product
        obj.qty = qty
        obj.uom = uom
        obj.subtotal = subtotal
        obj.save()
    return redirect('/po/addpo')


def getponum():
    num = random.randint(1, 100000000)
    return num


def deletechilditm(request, id):
    obj = poItems.objects.get(id=id)
    obj.delete()
    return redirect('/po/addpo')


def editchildqty(request, id):
    if request.method == 'POST':
        obj = poItems.objects.get(id=id)
        obj.qty = float(request.POST.get('eqty'))
        obj.subtotal = obj.qty * obj.po_items.price
        obj.save()
    return redirect('/po/addpo')

def sendemailToVendor(request,id):
    po = Po.objects.get(id=id)
    email = EmailMessage(
        'Simply Billing Purchase Order', 'PFA purchase order. Thanks', 'test@gmail.com',
        [po.vendor.email])
    email.attach_file('./files/' + po.ponum + '.xlsx')
    email.send(fail_silently=True)
    return redirect('/po/polist')

def donePo(request, id):
    if request.method == 'POST':
        PO = Po.objects.get(id=id)
        s = request.POST.get('status')
        taxp = float(request.POST.get('tax'))
        discp = float(request.POST.get('disc'))
        items = poItems.objects.filter(podetails=PO).filter(ponum=PO.ponum)
        total,calamt = 0,0
        ordt = potracker.objects.filter(podetail=PO).filter(po_num=PO.ponum)
        if ordt.exists():
            for i in items:
                calamt = i.subtotal + calamt
                total = ((calamt * taxp) / 100) + calamt
            ordt[0].total = total
            ordt[0].tax = taxp
            ordt[0].ponum = PO.ponum
            ordt[0].disc = discp
            ordt[0].status = s
            ordt[0].save()
        else:
            for i in items:
                calamt = i.subtotal+calamt
                total = ((calamt * taxp) / 100) + calamt
            print(total)
            obj = potracker()
            obj.status = s
            obj.tax = taxp
            obj.disc = discp
            obj.pototal = total
            obj.po_num = PO.ponum
            obj.podetail = PO
            obj.save()
        filename = './files/' + PO.ponum + '.xlsx'
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Purchase'
        sheet.merge_cells('A1:B2')
        cell1 = sheet.cell(row=1, column=1)
        cell1.value = 'Purchase Order'
        c1 = sheet.cell(row=4, column=1)
        c1.value = 'PO Number'
        c1.font.copy(bold=True)
        sheet.cell(row=4, column=2).value = 'Date'
        sheet.cell(row=4, column=3).value = 'Vendor'
        sheet.cell(row=4, column=4).value = 'Shipping Adderss'
        sheet.cell(row=4, column=5).value = 'Po Required Date'
        sheet.cell(row=5, column=1).value = PO.ponum
        sheet.cell(row=5, column=2).value = PO.date
        sheet.cell(row=5, column=3).value = PO.vendor.name
        sheet.cell(row=5, column=4).value = PO.shipto + PO.shipcity + PO.shipstate + PO.shippincode
        sheet.cell(row=5, column=5).value = PO.poreqbydate
        sheet.cell(row=8, column=1).value = 'Item'
        sheet.cell(row=8, column=2).value = 'Qty'
        sheet.cell(row=8, column=3).value = 'Unit Price'
        sheet.cell(row=8, column=4).value = 'Uom'
        sheet.cell(row=8, column=5).value = 'Sub Total'
        l2 = []
        for i in items:
            l1 = []
            l1.append(i.po_items.Name)
            l1.append(i.qty)
            l1.append(i.po_items.price)
            l1.append(i.uom.name)
            l1.append(i.subtotal)
            l2.append(l1)
        for itm in l2:
            sheet.append(itm)
        sheet.cell(row=9+len(l2), column=4).value = 'Total'
        sheet.cell(row=9+len(l2), column=5).value = total
        wb.save(filename)
    return redirect('/po/addpo')
